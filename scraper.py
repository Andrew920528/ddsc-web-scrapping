#!/usr/bin/env python
# coding: utf-8

import pandas as pd
import pprint
import requests
from bs4 import BeautifulSoup
import math
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import json
import os
from openpyxl import load_workbook
import argparse


def load_excel(excelPath, verbose = False):
    """ Format excel sheet to the desired format
    | Library | Vulnerabilities | Date | Url |
    | ------- | --------------- | ---- | --- |

    Args:
        excelPath (_type_): _path to excel file_
        verbose (bool, optional): _prints debug message or not_. Defaults to False.
    """
    
    df = pd.read_excel(excelPath,sheet_name=0)
    print(f"Loading {excelPath}")
    if verbose:
        print(f"{df.shape[0]} rows, {df.shape[1]}  columns")
    
    # Convert to a list of dictionary to iterate through
    # Assuming format is consistent. We're interested in Column G - L (7 - 12)
    cropped = df.iloc[1:, 5:12] # 0 based, first row is header

    libraries = cropped.iloc[:, 0]
    vulnerabilities = cropped.iloc[:,1]
    date= cropped.iloc[:, 3]
    url = cropped.iloc[:, 5]

    dependencies = []

    for i in range(libraries.shape[0]):
        dependencies.append({
            "library": libraries.iloc[i],
            "vulnerabilities": vulnerabilities.iloc[i],
            "date": date.iloc[i],
            "url": url.iloc[i]
        })
    if verbose:
        print(f"Here are the first few libraries:")
        pprint.pprint(dependencies[:5])
    print("Loaded excel file.")
    return dependencies



def scrape_data(dependencies, verbose):
    """
    Scrape the first row of the "versions" table from the given list of dependencies.
    The scraped data is then formatted and put into the same list of dictionaries.

    Args:
        dependencies (list): A list of dictionaries containing library name, url, 
            and other information.
        verbose (bool): Whether to print debug message or not.

    Returns:
        list: The list of dictionaries with scraped data.
    """
    first_rows_of_version_table = []
    failed_url = {}
    
    print(f"Start scrapping {len(dependencies)} libraries.")
    for i, dep in enumerate(dependencies[:]):
        if verbose:
            print(f"================== {i + 1} / {len(dependencies)} ==================")
        driver = webdriver.Chrome()
        url = dep["url"]
        start_time = time.perf_counter()
        try:
            if verbose:
                print(f"scraping: {url}")
            driver.get(url)

            # Find the first row of the "versions" table
            table = driver.find_element(By.CLASS_NAME, 'versions')
            thead = table.find_element(By.TAG_NAME, 'thead')
            tr = thead.find_element(By.TAG_NAME, "tr")
            ths = tr.find_elements(By.TAG_NAME, "th")
            
            headers= []
            for th in ths:
                headers.append(th.text)
            
            tbody = table.find_element(By.TAG_NAME, 'tbody')
            first_row = tbody.find_element(By.TAG_NAME, 'tr')
            cells = first_row.find_elements(By.TAG_NAME, 'td')

            # Extract the data from the first row
            rowData = {}
            for i, cell in enumerate(cells):
                try:
                    x = cell.text
                except:
                    x = ""
                rowData[headers[i]] =  x
            
            
        except Exception as e:
            # invalid url
            if verbose:
                print(f"Encountered error, pay attention to this url: {url}")
            rowData = {
                "Version" : "",
                "Vulnerabilities" : "",
                "Date" : ""
            }
            
            failed_url[i] = url
        finally:
            driver.close()
            # pass


        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        first_rows_of_version_table.append(rowData)
        if verbose:
            print(f"Elapsed time: {elapsed_time} seconds")
    
    driver.quit()
    
    print(f"Finished scrapping {len(dependencies)} libraries.")
    print(f"{len(failed_url)} urls failed during fetching:")
    print(failed_url)
    # Format the scraped data
    for i in range(len(dependencies[:])):
        row = first_rows_of_version_table[i]
        versionNumber = row["Version"]
        dateString = row["Date"]
        vulnerabilitiesString = row["Vulnerabilities"]
        
        formatedVersion = formatLibrary(dependencies[i]["library"], versionNumber)
        formatedVulnerability = formatVulnerability(vulnerabilitiesString)
        formatedDate = formatDate(dateString)
        
        dependencies[i]['date']= formatedDate
        dependencies[i]['library']= formatedVersion
        dependencies[i]['vulnerabilities']= formatedVulnerability

    print("Dependencies formatted.")
    if verbose:
        pprint.pprint(dependencies)
    
   
    return dependencies


def formatLibrary(libName, versionNumber):
    if (not libName or (type(libName) is float and math.isnan(libName))):
        return math.nan
    libNameList =libName.split('-')
    extension = libNameList[-1].split('.')[-1]
    return "".join(libNameList[:-1]) + "-" + versionNumber + "." + extension

def formatVulnerability(vulnerabilitiesString):
    if (not vulnerabilitiesString or (type(vulnerabilitiesString) is float and math.isnan(vulnerabilitiesString))):
        return math.nan
    return vulnerabilitiesString.split()[0]

def formatDate(dateString):
    if (not dateString or (type(dateString) is float and math.isnan(dateString))):
        return math.nan
    date_format = "%b %d, %Y"
    return datetime.strptime(dateString, date_format)


def write_json(dependencies, verbose):
    formatted_date = datetime.now().strftime('%Y%m%d')
    filename = f'{formatted_date}_dependencies.json'

    dep_copy = dependencies[:]
    for dep in dep_copy:
        for k, value in dep.items():
            if isinstance(value, float) and math.isnan(value):
                dep[k] = ""
            if isinstance(value, datetime):
                dep[k] = value.isoformat()


    # Writing the dictionary to a JSON file
    with open(filename, 'w') as file:
        json.dump(dep_copy, file, indent=4, ensure_ascii=False,)
    print(f"Saved {filename} to {filename}.")


def write_excel(dependencies, excelPath, verbose):
    # Create a new sheet in excel
    workbook = load_workbook(excelPath)
    sheet_to_duplicate = workbook.worksheets[0]

    # Duplicate the sheet
    new_sheet = workbook.copy_worksheet(sheet_to_duplicate)
    formatted_date = datetime.now().strftime('%Y%m%d') # Format the date as 'YYYYMMDD'
    new_sheet.title = "t" + str(formatted_date)

    workbook._sheets.insert(0, workbook._sheets.pop(workbook.sheetnames.index(new_sheet.title)))

    # Save the workbook
    workbook.save(excelPath)
    if verbose:
        print(f"Duplicated '{sheet_to_duplicate.title}' as '{new_sheet.title}' in '{excelPath}'")

    # Write data rows
    for row_idx, row_data in enumerate(dependencies, start=3):
        new_sheet.cell(row=row_idx, column=6, value=row_data.get('library'))
        new_sheet.cell(row=row_idx, column=7, value=row_data.get('vulnerabilities'))
        new_sheet.cell(row=row_idx, column=9, value=row_data.get('date'))
    workbook.save(excelPath)
    print(f"Saved '{new_sheet.title}' to '{excelPath}'")


def run(excelPath, verbose):
    print("Starting program...")
    dependencies = load_excel(excelPath, verbose)
    dependencies = scrape_data(dependencies, verbose)
    write_json(dependencies, verbose)
    write_excel(dependencies, excelPath, verbose)
    print("Program completed.")


curr_dir = os.path.dirname(os.path.abspath(__file__))
LOCAL_EXCEL_PATH = os.path.join(curr_dir,'永豐.保管系統.Lib元件版本清單.20240630.xlsx')
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Updates the excel file with the latest dependencies.")
    parser.add_argument("excel_path", type=str, help="Path to the excel file")
    parser.add_argument('-v', '--verbose', action='store_true', help="If true, prints out all messages. Else, prints out minimal messages and a progress bar.") 
    
    args = parser.parse_args()
    excel_path = args.excel_path
    verbose = args.verbose
    run(excel_path, verbose)
